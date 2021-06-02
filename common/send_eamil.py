import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.utils import parseaddr,formataddr
import datetime
from openpyxl import load_workbook
import os
import logging
import pandas as pd
import numpy as np
def handle_res(res):
    test_nums=0
    sucess_nums=0
    fail_nums=0

    print(type(res))
    if 'list' in str(type(res)):
        print("输入参数是list")
        for step in res:
            if step[2]==True:
                test_nums=test_nums+1
                sucess_nums=sucess_nums+1
            else:
                test_nums=test_nums+1
                fail_nums=fail_nums+1
    print("测试次数： "+str(test_nums))
    print("成功次数： "+str(sucess_nums))
    print("失败测试： "+str(fail_nums))
    return test_nums,sucess_nums,fail_nums

def handle_excel(filepath):
    wb=load_workbook(filepath)
    ws = wb.active
    row_max = ws.max_row
    con_max = ws.max_column

    print(row_max)
    print(con_max)


    data=pd.read_excel(filepath,sheet_name=0,header=1)
    print(data)
    print(type(data))
    # print(data.index)
    # print(data.columns)

    headers_names=['isrun','id','ispass','model','name','step1','ispass1','step2','ispass2','step3','ispass3','step4','ispass4','step5','ispass5','step6','ispass6','step7','ispass7','step8','ispass8','step9','ispass9','step10','ispass10']
    data2=data.loc[1:,headers_names]
    print(data2)
    res=[]
    for index,row in data2.iterrows():
        res_tmp=[]
        col=row[headers_names]
        header_steps=['step1','ispass1','step2','ispass2','step3','ispass3','step4','ispass4','step5','ispass5','step6','ispass6','step7','ispass7','step8','ispass8','step9','ispass9','step10','ispass10']
        print(col)
        print(type(col))
        print(col['name'])
        print(type(col['name']))
        print(col['isrun'])
        print(type(col['isrun']))
        if pd.isna(col['isrun']):
            print("这个是nan,说明是执行过的用例")
            print(col['ispass'])
            print(type(col['ispass']))
            if col['ispass']==True:
                print("这条用例通过了，只需要保存id，name,ispass即可")
                res_tmp.append(col['id'])
                res_tmp.append(col['name'])
                res_tmp.append(col['ispass'])
                res.append(res_tmp)
            else:
                print("执行失败了，需要保存id,name,ispass，失败的接口")
                res_tmp.append(col['id'])
                res_tmp.append(col['name'])
                res_tmp.append(col['ispass'])
                for i in range(1,11):
                    print(i)
                    step_name="step"+str(i)
                    step_ispass="ispass"+str(i)
                    if col[step_ispass]!=True:
                        print("该接口失败了")
                        stepname=col[step_name]
                        res_tmp.append(stepname)
                        break
                res.append(res_tmp)
        else:
            pass
            #未执行的就不要统计了
            # res_tmp.append(col['name'])
            # res_tmp.append("没有执行")
            # res.append(res_tmp)

        print("----------------------")
    print("++++++++++++++++")
    print(res)
    return res





def my_send_email(msg_to,file_path):
    logging.info("开始发送邮件")
    now_time = datetime.datetime.now()
    year = now_time.year
    month = now_time.month
    day = now_time.day
    mytime = str(year) + " 年 " + str(month) + " 月 " + str(day) + " 日 "
    msg_from = '1508691067@qq.com'  # 发送方邮箱
    passwd = 'fgaplzfksqsihdbe'

    subject = '接口自动化测试结果'

    res=handle_excel(file_path)
    test_nums,sucess_nums,fail_nums=handle_res(res)

    Success_rate=float(sucess_nums/test_nums)
    #print(Success_rate)
    rata = '%.2f%%' % (Success_rate * 100)
    logging.info("测试成功率： "+str(rata))

    content_base='''
                        <html>
                        <body>
                            <h1 align="center">接口自动化测试结果</h1>
                            <blockquote><p align="center"><strong><font size="4">测试用例数: {test_times}</font></strong></p></blockquote>
                            <blockquote><p align="center"><strong><font size="4">成功用例数: {sucess_nums}</font></strong></p></blockquote>
                            <blockquote><p align="center"><strong><font size="4">失败用例数: {fail_nums}</font></strong></p></blockquote>
                             <blockquote><p align="center"><strong><font size="4">成功率: {rata}</font></strong></p></blockquote>
                            <blockquote><p align="center"><strong><font size="4">附件是详细结果,请查收！</font></strong></p></blockquote>
                            <hr/>
                        <body>
                        <html>
                        '''.format(test_times=test_nums, sucess_nums=sucess_nums,fail_nums=fail_nums,rata=rata)

    content=content_base
    for step in res:
        if step[2]==True:
            content_tmp='''
                            <html>
                            <body>
                                <blockquote><p style="color:green;"><strong><font size="4">{id} : {step_name}  : {result}  </font></strong></p></blockquote>
                                <hr/>
                            <body>
                            <html>
                            '''.format(id=step[0], step_name=step[1], result=step[2])
            content=content+content_tmp
        else:
            #失败的html格式
            content_tmp = '''
                                <html>
                                <body>
                                    <blockquote><p style="color:red;"><strong><font size="4">{id} : {step_name}  :  {result} : 失败的接口是： {fial_name}</font></strong></p></blockquote>
                                    <hr/>
                                <body>
                                <html>
                                '''.format(id=step[0], step_name=step[1], result=step[2],fial_name=step[3])
            content=content+content_tmp

    content_tail='''
                <html>
                <body>
                    <p align="right"><font size="4">{mytime}</font></p>
                <body>
                <html>
                        '''.format(mytime=mytime)
    #构建Html对象
    content=content+content_tail
    # 构造要发送的内容格式
    content1 = '''
                        <html>
                        <body>
                            <h1 align="center">智能音箱唤醒成功率测试结果</h1>
                            <p><strong>您好：</strong></p>
                            <blockquote><p><strong>测试次数: {test_times}</strong></p></blockquote>
                            <blockquote><p><strong>唤醒成功次数: {sucess_nums}</strong></p></blockquote>
                            <blockquote><p><strong>唤醒成功率: {rata}</strong></p></blockquote>
                            <blockquote><p><strong>附件是语音唤醒测试结果,请查收！</strong></p></blockquote>


                            <p align="right">{mytime}</p>
                        <body>
                        <html>
                        '''.format(test_times=test_nums, sucess_nums=sucess_nums,rata=rata,mytime=mytime)
    #构建Html对象
    content2='''
                        <html>
                       
                        <body>
                            <h1 align="center">智能音箱唤醒成功率测试结果</h1>
                            <p><strong>您好：</strong></p>
                            <blockquote><p align="center" style="color:red;"><strong>测试次数: {test_times}</strong></p></blockquote>
                            <hr/>
                            <blockquote><p><strong>唤醒成功次数: {sucess_nums}</strong></p></blockquote>
                            <hr/>
                            <blockquote><p><strong>唤醒成功率: {rata}</strong></p></blockquote>
                            <hr/>
                            <blockquote><p><strong>附件是语音唤醒测试结果,请查收！</strong></p></blockquote>


                            <p align="right">{mytime}</p>
                        <body>
                        <html>
                        '''.format(test_times=test_nums, sucess_nums=sucess_nums,rata=rata,mytime=mytime)

    # content=content_base+content1+content2
    msg=MIMEMultipart()
    msg.attach(MIMEText(content,'html','utf-8'))

    att1=MIMEText(open(file_path,'rb').read(),'base64','utf-8')
    att1['Content-Type']='application/octet-stream'
    #file_base_path = os.path.dirname(file_path)  # 获取路径
    file_base_name = os.path.basename(file_path)  # 获取文件名称
    att1['Content-Disposition'] = 'attachment;filename=' + file_base_name
    msg.attach(att1)

    # 放入邮件主题
    msg['Subject'] = subject

    # 放入发件人,这是展示在邮件里面的，和时间的发件人没有关系
    msg['From'] = msg_from

    try:
        # 通过ssl方式发送，服务器地址，端口
        s = smtplib.SMTP_SSL("smtp.qq.com", 465)
        # 登录邮箱
        s.login(msg_from, passwd)
        s.sendmail(msg_from, msg_to, msg.as_string())
        print("邮件发送成功")
        logging.info("邮件发送成功")
    except Exception as e:
        # print(e)
        logging.error("发送邮件失败")
        logging.error(e)
    finally:
        logging.info(str(msg_to)+"已发送")
        logging.info("邮件发送结束")
        s.quit()




if __name__=="__main__":
    # Success_rate = 1 - float(1 / 18)
    # print(Success_rate)
    # rata='%.2f%%'%(Success_rate*100)
    # print(rata)
    # msg_to = ['1508691067@qq.com','wei_wei1992@yeah.net']
    # Success_rate = 1 - float(1 / 18)
    # print(Success_rate)
    # rata='%.2f%%'%(Success_rate*100)
    # print(rata)
    #msg_to = ['1508691067@qq.com','wei_wei1992@yeah.net']
    msg_to = ['1508691067@qq.com']
    file_path1='D:\\Python_Project\\API_Auto_Test\\result\\result_2021_04_30_09_59_08.xlsx'
    # test_times=5
    # sucess_nums=1
    my_send_email(msg_to,file_path1)
    # res=handle_excel(file_path)
    # test_nums, sucess_nums, fail_nums=handle_res(res)