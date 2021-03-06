import openpyxl
from openpyxl import load_workbook
import os
import time
import sys
import logging

from datetime import datetime
import time
from openpyxl.styles import PatternFill
from openpyxl.styles import Color,Font,Alignment
# import logging.config
# CON_LOG="../config/log.conf"
# logging.config.fileConfig(CON_LOG)
# logging=logging.getLogger()
# logging.info("xxxxxxxxxxxxxxxxxxxxxxx")
Excel_Headers = {
            "isrun": 1,
            "id": 2,
            "ispass":3,
            "model": 4,
            "name": 5,

            "step1": 6,
            "url1":7,
            "body1":8,
            "methond1":9,
            "version1":10,
            "verify1":11,
            "check1":12,
            "get_value1":13,
            "real_response1":14,
            "ispass1":15,
            "fail_reason1":16,


            "step2": 17,
            "url2": 18,
            "body2":19,
            "methond2": 20,
            "version2": 21,
            "verify2": 22,
            "check2":23,
            "get_value2":24,
            "real_response2":25,
            "ispass2":26,
            "fail_reason2":27,


            "step3": 28,
            "url3": 29,
            "body3": 30,
            "methond3": 31,
            "version3": 32,
            "verify3": 33,
            "check3":34,
            "get_value3":35,
            "real_response3":36,
            "ispass3":37,
            "fail_reason3":38,



            "step4": 39,
            "url4": 40,
            "body4": 41,
            "methond4": 42,
            "version4": 43,
            "verify4": 44,
            "check4":45,
            "get_value4":46,
            "real_response4":47,
            "ispass4":48,
            "fail_reason4":49,


            "step5": 50,
            "url5":51,
            "body5":52,
            "methond5":53,
            "version5":54,
            "verify5":55,
            "check5":56,
            "get_value5":57,
            "real_response5":58,
            "ispass5":59,
            "fail_reason5":60,


            "step6": 61,
            "url6":62,
            "body6":63,
            "methond6":64,
            "version6":65,
            "verify6":66,
            "check6":67,
            "get_value6":68,
            "real_response6":69,
            "ispass6":70,
            "fail_reason6":71,


            "step7":72 ,
            "url7":73,
            "body7":74,
            "methond7":75,
            "version7":76,
            "verify7":77,
            "check7":78,
            "get_value7":79,
            "real_response7":80,
            "ispass7":81,
            "fail_reason7":82,

            "step8":83 ,
            "url8":84,
            "body8":85,
            "methond8":86,
            "version8":87,
            "verify8":88,
            "check8":89,
            "get_value8":90,
            "real_response8":91,
            "ispass8":92,
            "fail_reason8":93,

            "step9":94 ,
            "url9":95,
            "body9":96,
            "methond9":97,
            "version9":98,
            "verify9":99,
            "check9":100,
            "get_value9":101,
            "real_response9":102,
            "ispass9":103,
            "fail_reason9":104,

            "step10": 105,
            "url10": 106,
            "body10": 107,
            "methond10": 108,
            "version10": 109,
            "verify10": 110,
            "check10": 111,
            "get_value10": 112,
            "real_response10": 113,
            "ispass10": 114,
            "fail_reason10": 115,

            "step11": 116,
            "url11": 117,
            "body11": 118,
            "methond11": 119,
            "version11": 120,
            "verify11": 121,
            "check11": 122,
            "get_value11": 123,
            "real_response11": 124,
            "ispass11": 125,
            "fail_reason11": 126,

            "step12": 127,
            "url12": 128,
            "body12": 129,
            "methond12": 130,
            "version12": 131,
            "verify12": 132,
            "check12": 133,
            "get_value12": 134,
            "real_response12": 135,
            "ispass12": 136,
            "fail_reason12": 137,

            "step13": 138,
            "url13": 139,
            "body13": 140,
            "methond13": 141,
            "version13": 142,
            "verify13": 143,
            "check13": 144,
            "get_value13": 145,
            "real_response13": 146,
            "ispass13": 147,
            "fail_reason13": 148,



        }
class Operate_Excel():
    def __init__(self,casefile):

        self.casefile=casefile
        # self.resultfile=self.creat_excel()

    def creat_excel(self):
        dt=datetime.now()
        now_time = dt.strftime('%Y_%m_%d_%H_%M_%S')
        my_path=os.path.abspath(os.getcwd())
        filename=my_path+'/result/result_%s.xlsx' %(now_time)

        wb=openpyxl.Workbook()
        mysheet=wb.active
        mysheet.merge_cells('A1:CD1')
        mysheet.cell(row=1,column=1,value="???????????????????????????")
        mysheet.row_dimensions[1].height = 25

        # ?????????????????????
        # ????????????????????????
        mycell = mysheet['A1']
        mycell.font = Font(name=u'??????', bold=True)
        mycell.alignment = Alignment(horizontal='center', vertical='center')

        # result_head = ['??????', '??????', '??????????????????', '?????????????????????', 'audio????????????', '??????ID']
        for i, item in enumerate(Excel_Headers):
            #print(i, item)
            mysheet.cell(row=2, column=i + 1, value=item).alignment = Alignment(horizontal='center', vertical='center')

        mysheet.title = "????????????"
        # mysheet.row_dimensions[3].height=25  #????????????,?????????3????????????

        wb.save(filename)
        # print(filename)
        logging.info("??????excel??????????????? " + str(filename))
        return filename

    def handle_casedata(self):
        wb=load_workbook(self.casefile)
        ws=wb.active
        row_max=ws.max_row
        con_max=ws.max_column
        con_titles=[]
        Cases=[]
        #????????????
        logging.info("????????????")
        for j in range(1, row_max + 1):
            tmp_case = {}
            for head in Excel_Headers:
                tmp_case[head] = ws.cell(j + 1, Excel_Headers[head]).value
            if tmp_case['id'] == None:
                # ??????id????????????
                break
            Cases.append(tmp_case)
        # logging.info("?????????????????? ")
        #logging.info(Cases)
        return Cases
    def save_result(self,resultfile,row,data):
        #content?????????????????????list
        #????????????????????????????????????????????????????????????
        wb=load_workbook(resultfile)
        ws=wb.active
        #columns=ws.max_column #??????
        columndata=[]
        data_len=len(data)
        for i in range(1,data_len+1):
            cellvalue=ws.cell(row=row,column=i).value
            ws.cell(row=row,column=i).value=data[i-1]
            #columndata.append(cellvalue)
        wb.save(resultfile)
        return columndata

def save_result(resultfile,row,data):
    #content?????????????????????list
    #????????????????????????????????????????????????????????????
    wb=load_workbook(resultfile)
    ws=wb.active

    #??????
    fill_fail = openpyxl.styles.PatternFill("solid", fgColor="FF0000")

    fill_pass = openpyxl.styles.PatternFill("solid", fgColor="00FF00")

    #columns=ws.max_column #??????
    columndata=[]
    data_len=len(data)
    flag=True
    for i in range(1,data_len+1):
        #cellvalue=ws.cell(row=row,column=i).value

        '''
        if str(v)=="False":
        ws.cell(row=row_max+1, column=i).fill=fill_fail
        '''
        if data[i-1]==None:
            # print("XXXXXXXXXXXXXXXXXX")
            # print(data[i-1])
            tmp=''
            ws.cell(row=row, column=i).value = tmp
        else:
            ws.cell(row=row,column=i).value=str(data[i-1])
            if str(data[i-1])=="False":
                flag=False
                ws.cell(row=row, column=i).fill=fill_fail
    if flag:
        #pass
        ws.cell(row=row, column=3).value = str(flag)
        ws.cell(row=row, column=3).fill = fill_pass

        # ws.cell(row=row,column=3).value=str(flag)
        # ws.cell(row=row,column=3).fill=fill_pass
    else:
        ws.cell(row=row, column=3).value = str(flag)
        ws.cell(row=row, column=3).fill = fill_fail
            #columndata.append(cellvalue)
    wb.save(resultfile)

def save_result_norun(resultfile,row,data):
    #??????????????????????????????
    #content?????????????????????list
    #????????????????????????????????????????????????????????????
    wb=load_workbook(resultfile)
    ws=wb.active

    #??????
    fill_fail = openpyxl.styles.PatternFill("solid", fgColor="FF0000")

    fill_pass = openpyxl.styles.PatternFill("solid", fgColor="00FF00")

    #columns=ws.max_column #??????
    columndata=[]
    data_len=len(data)
    flag=True
    for i in range(1,data_len+1):
        #cellvalue=ws.cell(row=row,column=i).value

        '''
        if str(v)=="False":
        ws.cell(row=row_max+1, column=i).fill=fill_fail
        '''
        if data[i-1]==None:
            # print("XXXXXXXXXXXXXXXXXX")
            # print(data[i-1])
            tmp=''
            ws.cell(row=row, column=i).value = tmp
        else:
            ws.cell(row=row,column=i).value=str(data[i-1])
            if str(data[i-1])=="False":
                flag=False
                ws.cell(row=row, column=i).fill=fill_fail
    if flag:
        pass
        # ws.cell(row=row, column=3).value = str(flag)
        # ws.cell(row=row, column=3).fill = fill_pass

        # ws.cell(row=row,column=3).value=str(flag)
        # ws.cell(row=row,column=3).fill=fill_pass
    else:
        ws.cell(row=row, column=3).value = str(flag)
        ws.cell(row=row, column=3).fill = fill_fail
            #columndata.append(cellvalue)
    wb.save(resultfile)

if __name__=="__main__":
    #OpExcel=Operate_Excel()
    file="D:\\Python_Project\\API_Auto_Test\\result\\result_2021_04_12_15_40_55.xlsx"
    OpExcel=Operate_Excel(file)
    data=['nihao','hhh','dfaf','afadf','afadfe']
    res=OpExcel.save_result(file,3,data)
    print(res)
    print(len(res))




